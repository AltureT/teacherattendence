import openpyxl
import pandas as pd


class User:
    def __init__(self, file_path):
        self.filename = file_path

        # 文件导入从周日或者周一开始，否则第一周情况会出现问题
        self.df = pd.read_excel(self.filename, sheet_name='每日统计')

        # ---------数据处理------------
        # 重设标题索引，丢弃无用行
        columns = self.df[1:2].values.tolist()
        self.df.columns = columns[0]
        self.df = self.df.drop([0, 1])

        # 重设行索引为userID，建立多层索引
        self.userid = self.df.loc[:, 'UserId'].values.tolist()
        self.name = self.df.loc[:, '姓名'].values.tolist()
        self.array = []
        self.array.append(self.userid)
        self.array.append(self.name)
        tuples = list(zip(*self.array))
        index = pd.MultiIndex.from_tuples(tuples, names=['userid', '姓名'])
        self.df.index = index
        self.df = self.df.fillna(value=-1)

        self.attendance = self.df.loc[:,
                          ['上班1打卡时间', '下班1打卡时间', '上班2打卡时间', '下班2打卡时间', '上班3打卡时间', '下班3打卡时间']].values.tolist()
        self.attendancedate = self.df.loc[:, '日期'].values.tolist()
        self.leave = self.df.loc[:, ['上班1打卡结果', '下班1打卡结果', '上班2打卡结果', '下班2打卡结果', '上班3打卡结果', '下班3打卡结果']].values.tolist()
        self.festival = self.df.loc[:, '班次'].values.tolist()
        self.userid = self.df.loc[:, 'UserId'].values.tolist()
        self.week = ('星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六')
        self.backschool = '星期日'

    # ---------每日次数统计,并添加到总数据------------
    # 获取样式[1，0，1]，三个值分别代表每日：合格考勤次数，迟到次数，严重迟到次数
    # attendancetimes = get_work_times(attendancedate, attendance, leave)

    def every_times_count(self, attendancetimes: list) -> None:
        x = []
        y = []
        z = []
        for i in range(len(attendancetimes)):
            x.append(attendancetimes[i][0])
            y.append(attendancetimes[i][1])
            z.append(attendancetimes[i][2])
        self.df['打卡次数'] = x
        self.df['缺30分钟以内次数'] = y
        self.df['缺30分钟以上次数'] = z

    # ---------计数功能实现------------
    def is_work(self, array: list) -> bool:
        '''

        :param array: ['请假','请假']
        :return: False
        '''
        if '请假' in array or '补卡审批通过' in array:
            return False
        return True

    def get_work_time(self, s1: str, s2: str) -> int:
        if s1 == -1 or s2 == -1:
            return 0
        hour1 = int(s1[0:2])
        minute1 = int(s1[3:5])
        hour2 = int(s2[0:2])
        minute2 = int(s2[3:5])
        if minute1 < minute2:
            minute1 += 60
            hour1 -= 1
        worktime = (hour1 - hour2) * 60 + minute1 - minute2
        return worktime

    def enough_work_time(self, array: list) -> bool:
        '''

        :param array: [06:52,11:45]
        :return: True
        '''

        if array[0] == -1 or array[1] == -1:
            return False
        # 工作时间按3小时（180）分钟计算,晚自习则根据最后打卡时间超过八点计算
        if self.get_work_time(array[1], array[0]) >= 180 or array[1][0:3] >= '20':
            return True

        return False

    def get_work_times(self) -> list:
        '''

        :param data: 21-09-01 星期三
        :param attendance: [06:52,11:45,13:52,17:14,0,0]
        :param leave:['请假','请假','请假','正常','缺卡','缺卡']
        :return:
        '''

        countresult = []
        for i in range(len(self.attendance)):
            c = 0
            late = 0
            severly = 0
            # 一天都是休息，或者工作日班次为'休息'，则代表法定节假日,也算老师打卡，除周六周日外
            if self.attendancedate[i][-3:] != '星期六' and \
                    self.attendancedate[i][-3:] != '星期日' and \
                    (self.leave[i].count('休息') >= 4 or self.festival[i] == '休息'):
                c = 2
            # 正常上班时，一日最多两次计数统计
            else:
                if 1 <= self.leave[i].count('请假') <= 3:
                    c = 1
                elif self.leave[i].count('请假') >= 4:
                    c = 2
                for j in range(0, len(self.attendance[0]), 2):
                    if c < 2:
                        worktime = self.get_work_time(self.attendance[i][j], self.attendance[i][j + 1])
                        # 请假,补卡也算正常打卡
                        if self.enough_work_time(self.attendance[i][j:j + 2]):
                            c = c + 1
                        elif 180 - worktime >= 30:
                            severly += 1
                        elif 0 < 180 - worktime < 30:
                            late += 1
            r = [c, late, severly]
            countresult.append(r)
        return countresult

    def limit_count(self, n: int) -> int:
        if n > 9:
            return 9
        return n

    def status(self, late: int, severly: int) -> str:
        if late + severly > 0:
            return '异常'
        return '正常'

    def every_week_summary_list(self, index: id, t: str, normal: list, late: list) -> list:
        x = self.limit_count(sum(normal))
        y = z = 0
        if x < 9:
            if sum(late) >= 9 - x:
                y = 9 - x
                z = 0
            else:
                y = sum(late)
                z = 9 - x - y

        return [self.name[index - 1], t, x, y, z, self.status(y, z)]

    # ---------每周次数统计------------

    def create_times_list(self, attendancetimes: list) -> dict:
        # 周日作为第一天，周日晚至周五晚，至少 4.5 天弹性坐班
        # 晚自习不需要满3小时算一次
        # 迟到早退 30 分钟(含)以上者计缺勤半天
        # 迟到早退 4 次(含合计)折 合缺勤半天 1 次
        # 连续旷工超过 15 个工作日或者一年内累计旷工超过 30 个工作 日的，按规定解除聘用合同
        # 未签到签退又未事前履行请假报告手续的视为缺勤
        # 每周因法定节假日导致不足5天，休息日视为两次

        # summary汇总记录每位教师每周打卡次数
        summary = {}

        # 缓存单位教师每周打卡次数
        normal = [0] * 7
        late = [0] * 7
        severly = [0] * 7
        start = 0
        end = 0
        t = str(self.attendancedate[start]) + ' -> ' + str(self.attendancedate[end])
        i = 0
        first = True
        while i < len(self.userid):
            id = str(self.userid[i])
            d = self.attendancedate[i][-3:]

            if id not in summary:
                summary[id] = []
                sameuser = False
            else:
                sameuser = True

            # 遇到第一人，或遇到同一人，且不在周六，则累计统计本周数据
            if first or sameuser and self.week.index(d) != 6:
                # 计入normal late severly，t数组，并增加长度
                first = False
                normal[self.week.index(d)] = attendancetimes[i][0]
                late[self.week.index(d)] = attendancetimes[i][1]
                severly[self.week.index(d)] = attendancetimes[i][2]
                end = i
                t = str(self.attendancedate[start]) + ' -> ' + str(self.attendancedate[end])

            # 遇到同一人，在周六，则统计本周数据，并保存结果，最后初始化每周临时数据
            elif sameuser and self.week.index(d) == 6:
                temp = self.every_week_summary_list(i, t, normal, late)
                summary[id].append(temp)
                normal = [0] * 7
                late = [0] * 7
                severly = [0] * 7
                start = i + 1

            # 遇到不同人，在周六，则统计本周数据给上一位用户，并保存结果，最后初始化每周临时数据，并更新新用户id
            elif not sameuser and self.week.index(d) == 6:
                temp = self.every_week_summary_list(i - 1, t, normal, late)
                summary[str(self.userid[i - 1])].append(temp)
                normal = [0] * 7
                late = [0] * 7
                severly = [0] * 7

                start = i + 1
                summary[id] = []

            # 遇到不同人，不在周六，有两种情况，一是数据只记录到中间，后续数据缺失，二是上一位用户周六统计完成，新用户开始，不需要处理
            # 假设上一位用户后面几天打卡完全，并保存结果给上一位用户，最后初始化每周临时数据，同时该天数据继续统计
            elif not first and not sameuser and self.week.index(d) != 6:
                # 第一种情况，如：
                # A，周三
                # A，周四；
                # B，周三；
                # B，周四
                if self.week.index(self.attendancedate[i - 1][-3:]) != 6:
                    for j in range(self.week.index(self.attendancedate[i - 1][-3:]) + 1, 6):
                        normal[j] = 2

                    temp = self.every_week_summary_list(i - 1, t, normal, late)
                    summary[str(self.userid[i - 1])].append(temp)
                    normal = [0] * 7
                    late = [0] * 7
                    severly = [0] * 7
                # 第二种情况
                else:
                    pass

                start = i
                i -= 1
                summary[id] = []

            else:
                print('遇到特殊情况，请检查数据·······')

            i += 1

        # 表格最后一组数据存储
        for j in range(self.week.index(self.attendancedate[i - 1][-3:]) + 1, 6):
            normal[j] = 2
        temp = self.every_week_summary_list(i - 1, t, normal, late)
        summary[str(self.userid[i - 1])].append(temp)

        return summary

    # 写入excel文件
    def write_excel(self, path: str, summary: dict):
        title = ['姓名', '日期', '正常打卡次数', '缺30分钟以内次数', '缺30分钟以上次数', '状态']

        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.create_sheet('按周统计汇总')

        for i in range(5):
            sheet.cell(row=1, column=i + 1, value=title[i])

        i = 1
        for key, val in summary.items():
            for j in range(len(val)):
                for k in range(len(val[j])):
                    sheet.cell(row=i + 1, column=k + 1, value=val[j][k])
                i += 1
        workbook.save(path)
        print('数据成功输出')
